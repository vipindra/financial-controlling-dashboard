"""
Excel management report generator.

Produces a formatted, multi-sheet Excel workbook that mirrors
the kind of monthly financial controlling report a finance team
would send to department heads and senior management.

Sheets:
    1. Executive Summary   - top-level KPIs and status overview
    2. Variance Analysis   - full YTD budget vs actuals by CC and category
    3. Monthly Trend       - actuals vs forecast month by month
    4. Cost Centre Health  - health grades and traffic light status
    5. Forecast Accuracy   - how well the rolling forecast performed
    6. Top Transactions    - largest individual spend items
    7. Category Breakdown  - total spend distribution by category
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from src.analysis import (
    build_executive_summary,
    analyse_variance_flags,
    analyse_cost_centre_health,
    analyse_forecast_accuracy,
    ExecutiveSummary,
)
from src.database import (
    get_monthly_trend,
    get_category_breakdown,
    get_top_transactions,
)

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "reports")

# ── Colour palette ─────────────────────────────────────────────────────────
C_NAVY       = "1B3A5C"
C_WHITE      = "FFFFFF"
C_LIGHT_BLUE = "D6E4F0"
C_HEADER_BG  = "1B3A5C"
C_ALT_ROW    = "F2F7FB"
C_GREEN      = "C6EFCE"
C_GREEN_FONT = "276221"
C_YELLOW     = "FFEB9C"
C_YELLOW_FNT = "9C6500"
C_RED        = "FFC7CE"
C_RED_FONT   = "9C0006"
C_ORANGE     = "FFE0B2"
C_ORANGE_FNT = "BF360C"


def _header_font(size=11, bold=True):
    return Font(name="Calibri", size=size, bold=bold, color=C_WHITE)

def _body_font(size=10, bold=False, color="111111"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")


def _write_header_row(ws, row: int, headers: list[str],
                       col_widths: list[int]) -> None:
    for col, (text, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font      = _header_font()
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_data_row(ws, row: int, values: list,
                     alternate: bool = False) -> None:
    bg = C_ALT_ROW if alternate else C_WHITE
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = _fill(bg)
        cell.border    = _border()
        cell.font      = _body_font()
        if isinstance(val, float):
            cell.alignment = _right()
        else:
            cell.alignment = _left()


def _variance_cell(ws, row: int, col: int, pct: float) -> None:
    """Colour a variance percentage cell by severity."""
    cell = ws.cell(row=row, column=col)
    cell.value     = pct / 100 if pct is not None else 0
    cell.number_format = "0.0%"
    cell.alignment = _center()
    cell.border    = _border()
    if pct is None:
        return
    if pct >= 15:
        cell.fill = _fill(C_RED)
        cell.font = _body_font(bold=True, color=C_RED_FONT)
    elif pct >= 7:
        cell.fill = _fill(C_YELLOW)
        cell.font = _body_font(color=C_YELLOW_FNT)
    elif pct <= -5:
        cell.fill = _fill(C_GREEN)
        cell.font = _body_font(color=C_GREEN_FONT)
    else:
        cell.fill = _fill(C_WHITE)
        cell.font = _body_font()


def _sheet_title(ws, title: str, subtitle: str = "") -> None:
    ws.row_dimensions[1].height = 28
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(name="Calibri", size=14, bold=True, color=C_NAVY)
    cell.alignment = _left()
    if subtitle:
        ws.row_dimensions[2].height = 16
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font      = Font(name="Calibri", size=9, color="666666", italic=True)
        sub.alignment = _left()


# ── Sheet builders ─────────────────────────────────────────────────────────

def _build_executive_summary(wb: Workbook, summary: ExecutiveSummary) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    _sheet_title(ws, "Financial Controlling Dashboard — Executive Summary",
                 f"Fiscal Year {summary.fiscal_year}  |  Generated {datetime.now().strftime('%d %b %Y')}")

    # KPI blocks starting at row 4
    kpis = [
        ("Total Annual Budget",        f"₹ {summary.total_annual_budget:,.2f}L",   C_LIGHT_BLUE),
        ("YTD Actuals",                f"₹ {summary.total_ytd_actuals:,.2f}L",     C_LIGHT_BLUE),
        ("YTD Variance",               f"₹ {summary.total_ytd_variance:,.2f}L",
         C_RED if summary.total_ytd_variance > 0 else C_GREEN),
        ("Variance %",                 f"{summary.total_ytd_variance_pct:+.1f}%",
         C_RED if summary.total_ytd_variance_pct > 0 else C_GREEN),
        ("Cost Centres Over Budget",   str(summary.cost_centres_over_budget),       C_ORANGE),
        ("Cost Centres On Track",      str(summary.cost_centres_on_track),          C_GREEN),
        ("Critical Alerts",            str(summary.categories_critical),            C_RED),
        ("Warning Alerts",             str(summary.categories_warning),             C_YELLOW),
        ("Avg Forecast Accuracy",      f"{summary.avg_forecast_accuracy_pct:.1f}%", C_LIGHT_BLUE),
        ("Months with Data",           str(summary.months_with_data),               C_LIGHT_BLUE),
        ("Largest Overspend",          summary.largest_overspend_cc,               C_ORANGE),
        ("Largest Underspend",         summary.largest_underspend_cc,              C_GREEN),
    ]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26

    for i, (label, value, colour) in enumerate(kpis):
        data_row = 4 + i
        lc = ws.cell(row=data_row, column=1, value=label)
        lc.font      = _body_font(bold=True)
        lc.fill      = _fill(colour)
        lc.border    = _border()
        lc.alignment = _left()

        vc = ws.cell(row=data_row, column=2, value=value)
        vc.font      = _body_font(bold=True, size=11)
        vc.fill      = _fill(colour)
        vc.border    = _border()
        vc.alignment = _center()


def _build_variance_analysis(wb: Workbook, fiscal_year: int,
                               db_path: str) -> None:
    from src.database import get_ytd_variance
    ws = wb.create_sheet("Variance Analysis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    _sheet_title(ws, "YTD Variance Analysis — Budget vs Actuals",
                 "Positive variance = overspend  |  Negative = underspend")

    headers = ["Cost Centre", "Owner", "Category",
               "Annual Budget (L)", "YTD Actuals (L)",
               "Variance (L)", "Variance %", "Status"]
    widths  = [28, 22, 26, 20, 20, 18, 13, 12]
    _write_header_row(ws, 4, headers, widths)

    rows = get_ytd_variance(fiscal_year, db_path)
    for i, r in enumerate(rows):
        data_row = 5 + i
        alt      = i % 2 == 1
        pct      = r["variance_pct"] or 0.0

        status = (
            "🔴 Critical" if pct >= 15 else
            "🟡 Warning"  if pct >= 7  else
            "🟢 On Track" if pct >= -5 else
            "🔵 Under"
        )

        _write_data_row(ws, data_row, [
            r["cost_centre"], r["owner"] or "", r["category"],
            round(r["annual_budget"], 2), round(r["ytd_actuals"], 2),
            round(r["variance"], 2), None, status
        ], alternate=alt)

        _variance_cell(ws, data_row, 7, pct)

        # Format currency columns
        for col in [4, 5, 6]:
            ws.cell(row=data_row, column=col).number_format = '#,##0.00'
            ws.cell(row=data_row, column=col).alignment     = _right()


def _build_monthly_trend(wb: Workbook, fiscal_year: int,
                          db_path: str) -> None:
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    _sheet_title(ws, "Monthly Actuals vs Rolling Forecast",
                 "All figures in INR Lakhs")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    headers = ["Month", "Actual Spend (L)", "Forecast Spend (L)",
               "Variance (L)", "Forecast Accuracy"]
    widths  = [12, 20, 22, 18, 20]
    _write_header_row(ws, 4, headers, widths)

    trend = get_monthly_trend(fiscal_year, db_path)
    for i, row in enumerate(trend):
        data_row = 5 + i
        alt = i % 2 == 1
        month_name = months[row["fiscal_month"] - 1]

        actual   = row["actual_spend"]
        forecast = row["forecast_spend"]
        variance = row["month_variance"]
        accuracy = max(0.0, 100 - abs(variance / max(forecast, 1) * 100))

        _write_data_row(ws, data_row, [
            month_name,
            round(actual, 2),
            round(forecast, 2),
            round(variance, 2),
            None
        ], alternate=alt)

        for col in [2, 3, 4]:
            ws.cell(row=data_row, column=col).number_format = '#,##0.00'
            ws.cell(row=data_row, column=col).alignment     = _right()

        acc_cell = ws.cell(row=data_row, column=5, value=accuracy / 100)
        acc_cell.number_format = "0.0%"
        acc_cell.alignment     = _center()
        acc_cell.border        = _border()
        if accuracy >= 95:
            acc_cell.fill = _fill(C_GREEN)
            acc_cell.font = _body_font(color=C_GREEN_FONT)
        elif accuracy >= 88:
            acc_cell.fill = _fill(C_YELLOW)
            acc_cell.font = _body_font(color=C_YELLOW_FNT)
        else:
            acc_cell.fill = _fill(C_RED)
            acc_cell.font = _body_font(color=C_RED_FONT)


def _build_cc_health(wb: Workbook, fiscal_year: int, db_path: str) -> None:
    ws = wb.create_sheet("Cost Centre Health")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    _sheet_title(ws, "Cost Centre Health Ratings",
                 "A = within 3% budget  |  F = over 25% variance")

    headers = ["Code", "Cost Centre", "Owner",
               "Total Budget (L)", "Total Actuals (L)",
               "Variance (L)", "Variance %",
               "Grade", "Over Categories", "Under Categories"]
    widths  = [10, 28, 22, 18, 18, 16, 13, 8, 18, 18]
    _write_header_row(ws, 4, headers, widths)

    health = analyse_cost_centre_health(fiscal_year, db_path)
    grade_colors = {
        "A": (C_GREEN,  C_GREEN_FONT),
        "B": (C_GREEN,  C_GREEN_FONT),
        "C": (C_YELLOW, C_YELLOW_FNT),
        "D": (C_ORANGE, C_ORANGE_FNT),
        "F": (C_RED,    C_RED_FONT),
    }

    for i, h in enumerate(health):
        data_row = 5 + i
        alt      = i % 2 == 1

        _write_data_row(ws, data_row, [
            h.code, h.name, h.owner,
            h.total_budget, h.total_actuals,
            h.total_variance, None,
            h.health_grade,
            h.categories_over, h.categories_under
        ], alternate=alt)

        for col in [4, 5, 6]:
            ws.cell(row=data_row, column=col).number_format = '#,##0.00'
            ws.cell(row=data_row, column=col).alignment     = _right()

        _variance_cell(ws, data_row, 7, h.variance_pct)

        grade_cell = ws.cell(row=data_row, column=8)
        fg, fc = grade_colors.get(h.health_grade, (C_WHITE, "111111"))
        grade_cell.fill      = _fill(fg)
        grade_cell.font      = Font(name="Calibri", size=11, bold=True, color=fc)
        grade_cell.alignment = _center()
        grade_cell.border    = _border()


def _build_top_transactions(wb: Workbook, fiscal_year: int,
                              db_path: str) -> None:
    ws = wb.create_sheet("Top Transactions")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    _sheet_title(ws, "Top 15 Transactions by Value",
                 "Largest individual expenditure items for spot-check and review")

    headers = ["Cost Centre", "Date", "Category", "Description", "Amount (L)"]
    widths  = [28, 14, 26, 38, 16]
    _write_header_row(ws, 4, headers, widths)

    txns = get_top_transactions(fiscal_year, db_path=db_path)
    for i, t in enumerate(txns):
        data_row = 5 + i
        alt = i % 2 == 1
        _write_data_row(ws, data_row, [
            t["cost_centre"], t["posting_date"],
            t["category"], t["description"],
            round(t["amount"], 2)
        ], alternate=alt)
        ws.cell(row=data_row, column=5).number_format = '#,##0.00'
        ws.cell(row=data_row, column=5).alignment     = _right()


def _build_category_breakdown(wb: Workbook, fiscal_year: int,
                                db_path: str) -> None:
    ws = wb.create_sheet("Category Breakdown")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    _sheet_title(ws, "Spend Distribution by Expense Category",
                 "Year-to-date totals across all cost centres")

    headers = ["Category", "Total Spend (L)", "Transactions", "% of Total"]
    widths  = [30, 20, 16, 14]
    _write_header_row(ws, 4, headers, widths)

    cats = get_category_breakdown(fiscal_year, db_path=db_path)
    for i, c in enumerate(cats):
        data_row = 5 + i
        alt = i % 2 == 1
        _write_data_row(ws, data_row, [
            c["category"],
            round(c["total_spend"], 2),
            c["transaction_count"],
            None
        ], alternate=alt)
        ws.cell(row=data_row, column=2).number_format = '#,##0.00'
        ws.cell(row=data_row, column=2).alignment     = _right()

        pct_cell = ws.cell(row=data_row, column=4,
                            value=c["pct_of_total"] / 100)
        pct_cell.number_format = "0.0%"
        pct_cell.alignment     = _center()
        pct_cell.border        = _border()
        pct_cell.fill          = _fill(C_ALT_ROW if alt else C_WHITE)
        pct_cell.font          = _body_font()


# ── Main export function ───────────────────────────────────────────────────

def generate_excel_report(fiscal_year: int, db_path: str) -> str:
    """
    Build and save the full Excel management report.
    Returns the file path of the saved workbook.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath  = os.path.join(REPORTS_DIR,
                              f"financial_controlling_report_{fiscal_year}_{timestamp}.xlsx")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    summary = build_executive_summary(fiscal_year, db_path)

    _build_executive_summary(wb, summary)
    _build_variance_analysis(wb, fiscal_year, db_path)
    _build_monthly_trend(wb, fiscal_year, db_path)
    _build_cc_health(wb, fiscal_year, db_path)
    _build_top_transactions(wb, fiscal_year, db_path)
    _build_category_breakdown(wb, fiscal_year, db_path)

    wb.save(filepath)
    return filepath
